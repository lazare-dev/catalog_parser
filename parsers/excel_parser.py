#!/usr/bin/env python3
"""
Excel parser for processing catalog data in Excel formats (.xlsx, .xls, .xlsm).
"""

import os
import logging
import openpyxl
import xlrd
from typing import List, Dict, Any, Tuple, Optional, Iterator
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException

from parsers.base_parser import BaseParser
from utils.error_handler import ParserError, FileError, log_error
from utils.manufacturer_detector import ManufacturerDetector
from utils.price_utils import PriceUtils
from config.settings import MAX_HEADER_ROWS

logger = logging.getLogger(__name__)

class ExcelParser(BaseParser):
    """
    Parser for Excel files. Handles both modern (.xlsx) and legacy (.xls) formats.
    """
    
    def __init__(self, filename: str, sheet_index: int = 0, sheet_name: Optional[str] = None):
        """
        Initialize the Excel parser.
        
        Args:
            filename: Path to the Excel file
            sheet_index: Index of the sheet to parse (default: 0)
            sheet_name: Name of the sheet to parse (overrides sheet_index if provided)
        """
        super().__init__(filename)
        self.sheet_index = sheet_index
        self.sheet_name = sheet_name
        self.workbook = None
        self.sheet = None
        self.sheet_names = []
        self.is_xlsx = filename.lower().endswith(('.xlsx', '.xlsm'))
        self.header_row_index = 0
        self.manufacturer_detector = ManufacturerDetector()
        
    def read_file(self) -> None:
        """
        Read and parse the Excel file.
        """
        logger.info(f"Reading Excel file: {self.basename}")
        
        try:
            if self.is_xlsx:
                self._read_xlsx_file()
            else:
                self._read_xls_file()
                
            logger.info(f"Successfully read Excel file with sheets: {self.sheet_names}")
            
        except Exception as e:
            error_msg = f"Error reading Excel file: {str(e)}"
            logger.error(error_msg)
            raise FileError(error_msg)
            
    def _read_xlsx_file(self) -> None:
        """
        Read modern Excel format (.xlsx, .xlsm) using openpyxl.
        """
        try:
            # Load workbook with data_only=True to get values instead of formulas
            self.workbook = openpyxl.load_workbook(self.filename, data_only=True)
            self.sheet_names = self.workbook.sheetnames
            
            # Select sheet based on name or index
            if self.sheet_name and self.sheet_name in self.sheet_names:
                self.sheet = self.workbook[self.sheet_name]
            else:
                # Use index with bounds checking
                if self.sheet_index >= len(self.sheet_names):
                    logger.warning(
                        f"Sheet index {self.sheet_index} out of bounds, using first sheet"
                    )
                    self.sheet_index = 0
                    
                self.sheet = self.workbook[self.sheet_names[self.sheet_index]]
                self.sheet_name = self.sheet_names[self.sheet_index]
                
            logger.info(f"Selected sheet: {self.sheet_name}")
            
            # Convert sheet data to list of rows
            self.data = []
            for row in self.sheet.iter_rows(values_only=True):
                # Skip completely empty rows
                if any(cell is not None and str(cell).strip() != "" for cell in row):
                    self.data.append(list(row))
                    
        except InvalidFileException:
            logger.warning("Invalid XLSX file, trying XLS format as fallback")
            self.is_xlsx = False
            self._read_xls_file()
        except Exception as e:
            raise FileError(f"Error reading XLSX file: {str(e)}")
            
    def _read_xls_file(self) -> None:
        """
        Read legacy Excel format (.xls) using xlrd.
        """
        try:
            # Open workbook
            self.workbook = xlrd.open_workbook(self.filename)
            self.sheet_names = self.workbook.sheet_names()
            
            # Select sheet based on name or index
            if self.sheet_name and self.sheet_name in self.sheet_names:
                self.sheet = self.workbook.sheet_by_name(self.sheet_name)
            else:
                # Use index with bounds checking
                if self.sheet_index >= len(self.sheet_names):
                    logger.warning(
                        f"Sheet index {self.sheet_index} out of bounds, using first sheet"
                    )
                    self.sheet_index = 0
                    
                self.sheet = self.workbook.sheet_by_index(self.sheet_index)
                self.sheet_name = self.sheet_names[self.sheet_index]
                
            logger.info(f"Selected sheet: {self.sheet_name}")
            
            # Convert sheet data to list of rows
            self.data = []
            for row_idx in range(self.sheet.nrows):
                row = self.sheet.row_values(row_idx)
                
                # Skip completely empty rows
                if any(str(cell).strip() != "" for cell in row):
                    self.data.append(row)
                    
        except Exception as e:
            raise FileError(f"Error reading XLS file: {str(e)}")
            
    def detect_headers(self) -> List[str]:
        """
        Detect and extract headers from the Excel file.
        
        Returns:
            List of header strings
        """
        logger.info("Detecting headers in Excel file")
        
        if not self.data:
            logger.warning("No data available to detect headers")
            return []
            
        # Find the most likely header row within the first MAX_HEADER_ROWS rows
        header_candidates = self.data[:min(MAX_HEADER_ROWS, len(self.data))]
        
        best_row = self._find_best_header_row(header_candidates)
        self.header_row_index = best_row
        
        # If no good header row found, use first row
        if best_row < 0:
            logger.warning("No good header row found, using first row")
            self.header_row_index = 0
            
        # Extract headers from identified row
        headers = self.data[self.header_row_index]
        
        # Clean up headers
        headers = [str(h).strip() if h is not None else f"Column{i+1}" 
                  for i, h in enumerate(headers)]
        
        # Remove empty headers
        for i, header in enumerate(headers):
            if not header:
                headers[i] = f"Column{i+1}"
                
        logger.info(f"Detected {len(headers)} headers: {headers}")
        return headers
        
    def _find_best_header_row(self, rows: List[List[Any]]) -> int:
        """
        Find the most likely header row from candidates.
        
        Args:
            rows: List of candidate header rows
            
        Returns:
            Index of the best header row, or -1 if none found
        """
        if not rows:
            return -1
            
        # Common header keywords to look for
        header_keywords = ['id', 'name', 'price', 'cost', 'sku', 'description', 'category', 
                          'brand', 'product', 'model', 'manufacturer', 'image', 'url']
                          
        best_score = -1
        best_row_index = -1
        
        for i, row in enumerate(rows):
            # Skip empty rows
            if not row or all(not cell for cell in row):
                continue
                
            score = 0
            
            # Check for typical header patterns
            
            # 1. String cells (non-numeric)
            string_ratio = sum(1 for cell in row if not self._is_numeric(cell)) / max(1, len(row))
            score += string_ratio * 3  # Weight towards string headers
            
            # 2. Keyword matches
            row_text = ' '.join(str(cell).lower() for cell in row if cell)
            keyword_matches = sum(1 for keyword in header_keywords if keyword in row_text)
            score += keyword_matches * 2
            
            # 3. Short text lengths (headers tend to be short)
            avg_len = sum(len(str(cell)) for cell in row if cell) / max(1, len(row))
            if avg_len < 30:  # Headers are usually shorter than data
                score += max(0, (30 - avg_len) / 10)
                
            # 4. Penalize rows with too many empty cells
            empty_ratio = sum(1 for cell in row if not cell) / max(1, len(row))
            score -= empty_ratio * 3
            
            if score > best_score:
                best_score = score
                best_row_index = i
                
        return best_row_index
        
    def _is_numeric(self, value: Any) -> bool:
        """
        Check if a value is numeric.
        
        Args:
            value: Value to check
            
        Returns:
            True if numeric, False otherwise
        """
        if value is None:
            return False
            
        # Check if it's already a number
        if isinstance(value, (int, float)):
            return True
            
        # Try to convert to float
        try:
            float(str(value).strip().replace(',', '.'))
            return True
        except (ValueError, TypeError):
            return False
            
    def _iterate_data_rows(self) -> Iterator[Dict[str, Any]]:
        """
        Iterate through data rows, converting each to a dictionary.
        
        Yields:
            Dictionary for each data row
        """
        if not self.data:
            return
            
        # Skip rows up to and including header row
        start_row = self.header_row_index + 1
        
        # Get headers
        headers = self.headers
        
        # Process each data row
        for row_index in range(start_row, len(self.data)):
            row = self.data[row_index]
            
            # Skip empty rows
            if not row or all(not cell for cell in row):
                continue
                
            # Create dictionary for row
            row_dict = {}
            
            # Map cells to headers
            for i, cell in enumerate(row):
                if i < len(headers):
                    header = headers[i]
                    row_dict[header] = cell
                else:
                    # Extra cells beyond header count
                    row_dict[f"ExtraColumn{i+1}"] = cell
                    
            yield row_dict
            
    def transform_data(self) -> List[Dict[str, Any]]:
        """
        Transform raw data to the target schema, with special handling for Excel files.
        
        Returns:
            List of dictionaries with standardized field names
        """
        transformed_data = super().transform_data()
        
        if not transformed_data:
            return []
            
        # Attempt to extract manufacturer from filename if not already set
        has_manufacturer = any(row.get('Manufacturer') for row in transformed_data)
        
        if not has_manufacturer:
            # Try to detect manufacturer from filename
            manufacturer = self.manufacturer_detector.detect_from_filename(self.filename)
            
            if manufacturer:
                logger.info(f"Setting manufacturer to '{manufacturer}' from filename")
                for row in transformed_data:
                    row['Manufacturer'] = manufacturer
        
        # Fix Excel specific date/time issues
        self._fix_excel_dates(transformed_data)
            
        # Process row-based price information
        self._process_row_based_prices(transformed_data)
        
        # Validate and normalize price fields
        for row in transformed_data:
            row = PriceUtils.validate_price_fields(row)
            
        return transformed_data
        
    def _fix_excel_dates(self, data: List[Dict[str, Any]]) -> None:
        """
        Fix Excel date/time values in the data.
        
        Args:
            data: List of data dictionaries to process
        """
        # Only needed for old XLS format
        if self.is_xlsx:
            return
            
        if not self.workbook:
            return
            
        # Process each row
        for row in data:
            for field, value in row.items():
                # Check if the value is a date/time number
                if isinstance(value, float) and 0 <= value <= 100000:  # Excel date range
                    try:
                        # Convert Excel date number to string date
                        date_tuple = xlrd.xldate_as_tuple(value, self.workbook.datemode)
                        # Format as ISO date string
                        if date_tuple[3:] == (0, 0, 0):  # No time component
                            row[field] = f"{date_tuple[0]}-{date_tuple[1]:02d}-{date_tuple[2]:02d}"
                        else:
                            row[field] = (
                                f"{date_tuple[0]}-{date_tuple[1]:02d}-{date_tuple[2]:02d} "
                                f"{date_tuple[3]:02d}:{date_tuple[4]:02d}:{date_tuple[5]:02d}"
                            )
                    except Exception as e:
                        # Not a valid date, leave as is
                        logger.debug(f"Error converting possible Excel date {value}: {str(e)}")
        
    def _process_row_based_prices(self, data: List[Dict[str, Any]]) -> None:
        """
        Process row-based price information that might be in description fields.
        
        Args:
            data: List of data dictionaries to process
        """
        # Check for price information in description fields
        for row in data:
            for field in ['Long Description', 'Short Description']:
                if field in row and row[field]:
                    # Try to extract prices from description
                    prices = PriceUtils.extract_prices_from_description(str(row[field]))
                    
                    # Apply extracted prices if found
                    for price_field, price_value in prices.items():
                        # Only set if not already present
                        if price_field not in row or row[price_field] is None:
                            row[price_field] = price_value